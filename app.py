from flask import Flask, request, send_file, render_template_string
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from io import BytesIO
import datetime
from collections import defaultdict
import pytz

app = Flask(__name__)

HTML_TEMPLATE = """
<!doctype html>
<html lang=\"ko\">
  <head>
    <meta charset=\"utf-8\">
    <title>배송리스트 → 발주서 변환기</title>
  </head>
  <body style=\"font-family: sans-serif; text-align: center; margin-top: 50px;\">
    <h1>📦 배송리스트 → 발주서 자동 변환</h1>
    <form action=\"/convert\" method=\"post\" enctype=\"multipart/form-data\">
      <input type=\"file\" name=\"file\" accept=\".xlsx\" required><br><br>
      <button type=\"submit\" style=\"font-size: 16px;\">발주서 만들기</button>
    </form>
  </body>
</html>
"""

unit_prices = {
    "14mm이상 400g": 17000,
    "14mm이상 600g": 24000,
    "14mm이상 1kg": 37000,
    "16mm이상 400g": 20000,
    "16mm이상 600g": 28000,
    "16mm이상 1kg": 44000,
    "18mm이상 400g": 23000,
    "18mm이상 600g": 32500,
    "18mm이상 1kg": 53000
}

def format_option_text(text):
    mapping = {
        "대(14mm~16mm)": "14mm이상",
        "특대(16mm~18mm)": "16mm이상",
        "왕특(18mm이상)": "18mm이상"
    }
    for key, val in mapping.items():
        if key in text:
            weight = text.split("_")[-1].strip()
            return f"{val} {weight}"
    return text

@app.route("/", methods=["GET"])
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route("/convert", methods=["POST"])
def convert():
    file = request.files['file']
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    input_wb = openpyxl.load_workbook(file)
    input_ws = input_wb.active

    col_map = {cell.value: idx for idx, cell in enumerate(input_ws[1])}

    ws = wb.create_sheet(title="발주서")
    headers = ["주문번호", "주문상품명", "상품모델", "수량", "수취인명", "수취인 우편번호",
               "수취인 주소", "수취인 전화번호", "수취인 이동통신", "배송메시지", "상품코드", "주문자명", "박스단위", "부피단위"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = []
    summary = defaultdict(int)
    for row in input_ws.iter_rows(min_row=2, values_only=True):
        option = row[col_map["등록옵션명"]]
        formatted = format_option_text(option)
        item = [
            row[col_map["주문번호"]],
            formatted,
            formatted,
            row[col_map["구매수(수량)"]],
            row[col_map["수취인이름"]],
            row[col_map["우편번호"]],
            row[col_map["수취인 주소"]],
            row[col_map["수취인전화번호"]],
            row[col_map["수취인전화번호"]],
            row[col_map["배송메세지"]],
            row[col_map["주문번호"]],
            row[col_map["구매자"]],
            1,
            60
        ]
        ws.append(item)
        rows.append(item)
        summary[formatted] += int(row[col_map["구매수(수량)"]])

    for col_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max_length + 5

    summary_ws = wb.create_sheet(title="발주 정리표")
    kst = datetime.datetime.now(pytz.timezone("Asia/Seoul"))
    today_kr = kst.strftime("%-m월 %-d일")
    summary_ws.append([f"{today_kr} 하입월드 발주"])
    summary_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    summary_ws.cell(row=1, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    summary_ws.cell(row=1, column=1).font = Font(bold=True)
    summary_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    summary_ws.append(["사이즈", "개수", "단가", "합계"])
    total_count = 0
    total_sum = 0
    for key in summary:
        count = summary[key]
        price = unit_prices.get(key, 0)
        total = count * price
        total_count += count
        total_sum += total
        summary_ws.append([key, count, price, total])

    # 합계 행 추가 및 노란색 채우기
    total_row = summary_ws.max_row + 1
    summary_ws.append(["", total_count, "", total_sum])
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    summary_ws.cell(row=total_row, column=2).fill = yellow_fill
    summary_ws.cell(row=total_row, column=4).fill = yellow_fill

    # 테두리 적용
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    for col in summary_ws.iter_cols(min_row=2, max_row=summary_ws.max_row):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        summary_ws.column_dimensions[col[0].column_letter].width = max_length + 5

    buffer = BytesIO()
    filename = f"{kst.strftime('%y%m%d')} 하입월드 발주서.xlsx"
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == "__main__":
    app.run(debug=True)
